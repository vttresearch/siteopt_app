import os
import sys
import json
import csv
import shutil
import uuid
import subprocess
import time
from pathlib import Path
from tempfile import TemporaryDirectory
import openpyxl
import spinedb_api.exception
from openpyxl.utils import range_boundaries
from spinedb_api import DatabaseMapping
from django.http import JsonResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from django.conf import settings
from django.core.cache import cache


IN_CONTAINER = os.environ.get("RUNNING_IN_CONTAINER", False)
WORK_DIR = "work_container" if IN_CONTAINER else "work"
CONFIG_FILE = "config.json"
INPUT_DATA_DIR = (Path(settings.BASE_DIR) / "siteopt_data").resolve()
EXAMPLE_INPUT_DATA_DIR = (Path(settings.BASE_DIR) / "siteopt_toolbox" / "example_data").resolve()
PROJECT_DATA_DIR = (Path(settings.BASE_DIR) / "siteopt_toolbox").resolve()
TEST_PROJECT_DATA_DIR = (Path(settings.BASE_DIR) / "test_spinetoolbox_project").resolve()
SERVER_CONFIG_PATH = Path(os.environ.get("SPINE_SERVER_CONFIG", Path(settings.BASE_DIR) / "server_config.txt")).resolve()
WORK_ROOT = Path(Path(settings.BASE_DIR) / WORK_DIR).resolve()
CONFIG_ROOT = Path(settings.BASE_DIR / "_config").resolve()
PYTHON_EXECUTABLE = sys.executable
INPUT_DATA_SQLITE_FILE = Path(".spinetoolbox", "items", "input_data", "elexia_input.sqlite")
_MOD_SCRIPT_NAME = "mod_script.py"


def get_input_data_path() -> str:
    return str(INPUT_DATA_DIR)


def get_example_input_data_path() -> str:
    return str(EXAMPLE_INPUT_DATA_DIR)


def get_project_data_path() -> str:
    return str(PROJECT_DATA_DIR)


def get_test_project_data_path() -> str:
    return str(TEST_PROJECT_DATA_DIR)


def get_config_file_dir(client_id) -> Path:
    return (CONFIG_ROOT / str(client_id)[0:6]).resolve()


def get_client_work_root(client_id: str) -> str:
    """Root directory containing all work folders for this client.
    Must be a path that BOTH backend and spine_engine containers can access."""
    return str((WORK_ROOT / str(client_id)[0:6]).resolve())


def list_results(project_path):
    root = Path(project_path) / ".spinetoolbox" / "items" / "extract_results" / "output"
    if not root.exists():
        return {}
    runs = {}
    for scenario_dir in root.iterdir():
        if not scenario_dir.is_dir():
            continue
        scenario_name = scenario_dir.name
        for run_dir in scenario_dir.iterdir():
            if not run_dir.is_dir():
                continue
            results_file = run_dir / "results.xlsx"
            if not results_file.exists():
                continue
            run_name = run_dir.name
            if run_name not in runs:
                runs[run_name] = []
            runs[run_name].append({
                "scenario": scenario_name,
                "run": run_name,
                "path": str(results_file)
            })
    for run_name in runs:
        runs[run_name].sort(key=lambda x: x["scenario"].lower())
    return dict(sorted(runs.items(), key=lambda x: x[0], reverse=True))


def list_projects_with_results(client_id):
    config = get_client_config(client_id)
    projects = config.get("work_folders", {})
    projects_with_results = []
    for name, path in projects.items():
        results_root = Path(path) / ".spinetoolbox" / "items" / "extract_results" / "output"
        if not results_root.exists():
            continue
        # check if any results.xlsx exists
        has_results = any(results_root.rglob("results.xlsx"))
        if has_results:
            projects_with_results.append({
                "name": name,
                "path": path
            })
    projects_with_results.sort(key=lambda x: x["name"].lower())
    return projects_with_results


@ensure_csrf_cookie
def health_check(request):
    """For polling the backend."""
    return JsonResponse({"status": "ok"})


def settings(request):
    client_id = request.COOKIES.get("client_id") or request.headers.get("X-Client-ID")
    print(f"Client {client_id} retrieving settings")
    new_client = False
    if not client_id:
        client_id = uuid.uuid4()
        new_client = True
    client_config = get_client_config(client_id)
    work_root = get_client_work_root(client_id)
    active_work_folders = {}
    # Edit work_folders so when in container, only the projects in the container work folder are returned
    for key, value in client_config["work_folders"].items():
        # key is project name, value is the path
        if value.startswith(work_root):
            active_work_folders[key] = value
    client_config["work_folders"] = active_work_folders
    print(f"client_config work_folders:{client_config['work_folders']}")
    response = JsonResponse({"success": True, "data": {"client_id": client_id, "configs": client_config}})
    if new_client:
        # httponly=False makes sure that we can access it from JavaScript
        response.set_cookie("client_id", client_id, httponly=False, samesite="Lax", max_age=31536000)  # 1 year
    return response


def get_client_config(client_id):
    """Returns config dict from the config file for given client id."""
    config_file_dir = get_config_file_dir(client_id)
    config_file_path = config_file_dir / CONFIG_FILE

    if not os.path.exists(str(config_file_dir)):
        make_dir(str(config_file_dir))
        print(f"Created directory:{str(config_file_dir)}")
        make_config_file(str(config_file_path))
    else:
        if not os.path.exists(str(config_file_path)):
            make_config_file(str(config_file_path))
    return read_config_file(str(config_file_path))


def remove_work_folder(client_id: str, work_folder_name: str):
    if not work_folder_name:
        return {"success": False, "error": "Missing work_folder"}
    config_fpath = get_config_file_dir(client_id) / CONFIG_FILE
    cfg = read_config_file(str(config_fpath))
    work_folders = cfg.get("work_folders", {})
    if work_folder_name not in work_folders:
        return {"success": False, "error": f"Unknown work folder: {work_folder_name}"}
    # Soft remove: only remove from config, keep files on disk
    work_folders.pop(work_folder_name, None)
    edit_config_file(config_fpath, {"work_folders": work_folders})
    return {"success": True, "data": {}}


def list_existing_work_folders(client_id: str):
    config_fpath = get_config_file_dir(client_id) / CONFIG_FILE
    print(f"config_fpath:{str(config_fpath)}")
    root = get_client_work_root(client_id)
    if not os.path.exists(root):
        return {"success": True, "data": []}
    cfg = read_config_file(str(config_fpath))
    active = cfg.get("work_folders", {})
    active_paths = set(os.path.abspath(p) for p in active.values())
    out = []
    for entry in os.listdir(root):
        p = os.path.join(root, entry)
        if not os.path.isdir(p):
            continue
        # only show projects not currently in view
        if os.path.abspath(p) in active_paths:
            continue
        # (optional) sanity check: looks like a Spine Toolbox project
        if not validate_project_path(p)["success"]:
            continue
        out.append({"name": entry, "path": p})
    out.sort(key=lambda x: x["name"].lower())
    return {"success": True, "data": out}


def add_existing_work_folder(client_id: str, name: str, path: str):
    if not name or not path:
        return {"success": False, "error": "Missing work_folder or path"}
    if not os.path.exists(path):
        return {"success": False, "error": f"Path does not exist: {path}"}
    config_fpath = get_config_file_dir(client_id) / CONFIG_FILE
    # Ensure this is under the client work root (prevents adding arbitrary dirs)
    cfg = read_config_file(str(config_fpath))
    # You can also pass client_id and check get_client_work_root(client_id)
    # For now, at least validate it is a project
    if not validate_project_path(path)["success"]:
        return {"success": False, "error": "Folder is not a valid Spine Toolbox project"}
    work_folders = cfg.get("work_folders", {})
    work_folders[name] = path
    edit_config_file(config_fpath, {"work_folders": work_folders})
    return {"success": True, "data": {}}


def fetch_scenarios_from_input_db(client_id, configs):
    db_key = configs["db_key"]
    client_root = Path(get_client_work_root(client_id))
    sqlite_fpath = (client_root / configs["work_folder"] / INPUT_DATA_SQLITE_FILE).resolve()
    if not os.path.exists(sqlite_fpath):
        return JsonResponse({"success": False, "error": f"SQLite file {sqlite_fpath} doesn't exist"})
    scenario_names = list()
    with DatabaseMapping("sqlite:///" + str(sqlite_fpath)) as db_map:
        scenarios = db_map.find_by_type(db_key)  # scenarios is a list-of-PublicItems
        for scenario in scenarios:
            scenario_names.append(scenario["name"])
    return JsonResponse({"success": True, "data": {"scenarios": scenario_names}})


def add_scenario(client_id, scenario_name, project_name):
    client_root = Path(get_client_work_root(client_id))
    sqlite_fpath = (client_root / project_name / INPUT_DATA_SQLITE_FILE).resolve()
    if not os.path.exists(sqlite_fpath):
        return JsonResponse({"success": False, "error": f"Input data SQlite file {sqlite_fpath} doesn't exist"})
    with DatabaseMapping("sqlite:///" + str(sqlite_fpath)) as db_map:
        # Add scenario item
        items, errors = db_map.add_items("scenario", {"name": scenario_name})
        if isinstance(errors, list) and len(errors) > 0:
            print(f"Something went wrong when adding scenarios. Errors: {errors}")
        try:
            db_map.commit_session(f"Add scenario {scenario_name}")
        except spinedb_api.exception.NothingToCommit:
            print("Nothing to commit")
        except spinedb_api.exception.SpineDBAPIError as e:
            print(f"Commit failed: {e}")
    return JsonResponse({"success": True, "data": {}})


def remove_scenario(client_id, scenario_name, project_name):
    client_root = Path(get_client_work_root(client_id))
    sqlite_fpath = (client_root / project_name / INPUT_DATA_SQLITE_FILE).resolve()
    if not os.path.exists(sqlite_fpath):
        return JsonResponse({"success": False, "error": f"Input data SQlite file {sqlite_fpath} doesn't exist"})
    with DatabaseMapping("sqlite:///" + str(sqlite_fpath)) as db_map:
        # Remove scenario item
        # scenario_table = db_map.mapped_table("scenario")
        # scenario_item = db_map.item(scenario_table, {})
        scenario = db_map.scenario(name=scenario_name)
        scenario.remove()
        try:
            db_map.commit_session(f"Remove scenario {scenario_name}")
        except spinedb_api.exception.NothingToCommit:
            print("Nothing to commit")
        except spinedb_api.exception.SpineDBAPIError as e:
            print(f"Commit failed: {e}")
    return JsonResponse({"success": True, "data": {}})


def delete_project(client_id, path):
    if not os.path.exists(path):
        # Return True because it's probably been deleted manually
        return JsonResponse({"success": True, "data": {}})
    # Remove path from disk
    try:
        shutil.rmtree(path)
    except OSError as e:
        return JsonResponse({"success": False, "error": f"Deleting project failed. [OSError] {e}"})
    # No need to edit config file because the project name and path have been removed from the config file
    # when the tab was closed.
    return JsonResponse({"success": True, "data": {}})


def replace_file(client_id, file_path, file_data):
    # filepath = data["fpath"]
    print(f"filepath:{file_path}")
    # Delete old file
    if os.path.exists(file_path):
        os.remove(file_path)

    with open(file_path, "wb+") as destination:
        for chunk in file_data.chunks():
            destination.write(chunk)

    return JsonResponse({"success": True, "data": {}})

    # Save new file
    # with open(target_path, "wb+") as destination:
    #    for chunk in uploaded_file.chunks():
    #         destination.write(chunk)
    # return JsonResponse({"status": "ok", "message": "File replaced successfully"})

    # file_data = data["formData"]
    # if not fileData:
    #     print("No file uploaded")
    #     return JsonResponse({"success": False, "error": "No file uploaded"})
    # print(f"file_data:{fileData}")
    # print(f"data:{data}")

# def replace_file(request):
#     if request.method == "POST":
#         uploaded_file = request.FILES.get("file")
#
#         if not uploaded_file:
#             return JsonResponse({"error": "No file uploaded"}, status=400)
#
#         # Path to the file on the server you want to replace
#         target_path = os.path.join(settings.MEDIA_ROOT, "myfile.xlsx")
#
#         # Delete old file if exists
#         if os.path.exists(target_path):
#             os.remove(target_path)
#
#         # Save new file
#         with open(target_path, "wb+") as destination:
#             for chunk in uploaded_file.chunks():
#                 destination.write(chunk)
#
#         return JsonResponse({"status": "ok", "message": "File replaced successfully"})
#
#     return JsonResponse({"error": "Invalid method"}, status=405)


@csrf_protect
def post(request, action):
    """Handles data posted by the frontend.
    Requires that the POST from frontend includes csrftoken cookie and
    'credentials: 'include'.
    Note: @csrf_protect decorator is needed for views that modify data (POST, PUT, DELETE)
    """
    client_id = request.COOKIES.get("client_id") or request.headers.get("X-Client-ID")
    if action != "upload_file":
        js = json.loads(request.body.decode("utf-8"))  # dict
        data = js["data"]
    if action == "make_work_folder":
        if "test_work_folder" in data.keys():
            print(f"[{client_id}] creating test project")
            json_response = make_test_work_folder(client_id, data["test_work_folder"])
        elif "work_folder_with_example_data" in data.keys():
            print(f"[{client_id}] creating SiteOpt project with example data")
            json_response = make_work_folder(client_id, data["work_folder_with_example_data"], use_example_data=True)
        else:
            print(f"[{client_id}] creating SiteOpt project")
            json_response = make_work_folder(client_id, data["work_folder"])
        return JsonResponse(json_response)
    elif action == "fetch_data":
        print(f"[{client_id}] fetching {data['full_path']}")
        response = fetch_data(data["full_path"])
        return JsonResponse(response)
    elif action == "execute":
        print(f"[{client_id}] executing {data['work_dir_name']}")
        return prepare_execution(client_id, data)
    elif action == "save_file":
        print(f"[{client_id}] saving {data['path']}")
        return JsonResponse(save_file(client_id, data["path"], data["filetype"], data["payloadData"], data["meta"]))
    elif action == "remove_work_folder":
        return JsonResponse(remove_work_folder(client_id, data["folder_name"]))
    elif action == "list_existing_work_folders":
        return JsonResponse(list_existing_work_folders(client_id))
    elif action == "add_existing_work_folder":
        return JsonResponse(add_existing_work_folder(client_id, data["name"], data["path"]))
    elif action == "fetch_input_db_data":
        if data["db_key"] == "scenario":
            print("Fetching scenarios")
            response = fetch_scenarios_from_input_db(client_id, data)
        else:
            print(f"Not implemented db_key:{data['db_key']}")
            response = JsonResponse({"success": False, "error": f"Fetching db_key {data['db_key']} not Implemented"})
        return response
    elif action == "add_scenario":
        print(f"Adding scenario {data['scenario_name']}")
        response = add_scenario(client_id, data["scenario_name"], data["work_folder"])
        return response
    elif action == "remove_scenario":
        print(f"Removing scenario {data['scenario_name']}")
        response = remove_scenario(client_id, data["scenario_name"], data["work_folder"])
        return response
    elif action == "list_results":
        project_name = data["project_name"]
        config = get_client_config(client_id)
        project_path = config["work_folders"].get(project_name)
        if not project_path:
            return JsonResponse({"success": False, "error": "Project not found"})
        results = list_results(project_path)
        return JsonResponse({"success": True, "data": results})
    elif action == "list_projects_with_results":
        projects = list_projects_with_results(client_id)
        return JsonResponse({"success": True, "data": projects})
    elif action == "delete_project":
        print(f"Deleting project {data['name']} path {data['path']}")
        response = delete_project(client_id, data["path"])
        return response
    elif action == "upload_file":
        file_data = request.FILES.get("file")
        file_path = request.POST.get("fpath")
        print(f"file_data:{file_data}")
        response = replace_file(client_id, file_path, file_data)
        return response
    else:
        print(f"Unknown action: {action}")
        return JsonResponse({"success": False, "error": f"No handler for action {action}"})


def prepare_execution(client_id, data):
    config = get_client_config(client_id)
    project_path = config["work_folders"][data["work_dir_name"]]
    if not os.path.exists(project_path):
        return JsonResponse({"success": False, "error": f"Path {project_path} does not exist"})
    job_id = str(uuid.uuid4())
    cache.set(
        job_id,
        {
            "path": project_path,
            "exec_items": data["executed_items"],
            "local": data["local_execution"],
            "scenarios": data["scenarios"]
        },
        timeout=300
    )
    return JsonResponse({"success": True, "data": {"job_id": job_id}})


def execute(request, job_id):

    def event_stream():
        timeout = 5
        start = time.time()
        job = cache.get(job_id)
        while not job and time.time() - start < timeout:
            time.sleep(0.1)
            job = cache.get(job_id)
        if not job:
            cache.delete(job_id)
            yield f"event: error\ndata: Execution failed. Job {job_id} not ready after {timeout}s.\n\n"
            return
        ppath = job["path"]
        items_to_execute = job["exec_items"]
        exec_locally = job["local"]
        project_path = Path(ppath).resolve()
        scenarios = job["scenarios"]
        temp_dir, script_path = _make_solve_model_mod_script(scenarios)
        print(f"Executing project {project_path} with scenarios {scenarios}")
        # Check that server config file exists if executing remotely.
        # TODO: These should be included in the execute request
        if not exec_locally and not SERVER_CONFIG_PATH.exists():
            cache.delete(job_id)
            yield (f"event: error\ndata: Execution failed. Config file "
                   f"(server_config.txt) for remote execution missing: {SERVER_CONFIG_PATH}\n\n")
            return
        if exec_locally:
            args = [
                PYTHON_EXECUTABLE,
                "-u", "-m",
                "spinetoolbox",
                "--mod-script",
                str(script_path),
                "--execute-only",
                str(project_path),
            ]
        else:
            args = [
                PYTHON_EXECUTABLE,
                "-u", "-m",
                "spinetoolbox",
                "--execute-only",
                "--execute-remotely", str(SERVER_CONFIG_PATH),
                str(project_path),
            ]
        item_args = [] if not items_to_execute else ["-s"] + items_to_execute
        if item_args:
            args += item_args
        try:
            proc = subprocess.Popen(args, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
            for line in iter(proc.stdout.readline, b""):
                line = line.decode("utf-8", "replace").strip()
                if line.startswith("Executing") and line.endswith("finished"):
                    # This intercepts the 'Executing ... finished' output line and sends an event instead
                    parsed_line = line.removeprefix("Executing").removesuffix("finished").strip()
                    # Strip project item type
                    parsed_line = parsed_line.removeprefix("Data Connection ")
                    parsed_line = parsed_line.removeprefix("Tool ")
                    parsed_line = parsed_line.removeprefix("Importer ")
                    parsed_line = parsed_line.removeprefix("Data Store ")
                    parsed_line = parsed_line.removeprefix("Merger ")
                    yield f"event: item_finished\ndata: {parsed_line}\n\n"
                else:
                    yield f"data: {line}\n\n"
            proc.stdout.close()
            proc_retval = proc.wait()
            cache.delete(job_id)
            # Notify frontend that execution is done (send process exit code)
            print("Execution finished")
            yield f"event: done\ndata: {proc_retval}\n\n"
        except OSError as e:
            cache.delete(job_id)
            print(f"Execution failed: [OSError]: {e}")
            yield f"event: error\ndata: Execution error: [OSError]: {e}\n\n"

    response = StreamingHttpResponse(event_stream(), content_type="text/event-stream")
    response["Cache-Control"] = "no-cache"
    return response


def _make_solve_model_mod_script(scenarios):
    """Writes project modification script for solve model to a temporary file.

    Args:
        scenarios (list of str): active scenario names

    Returns:
        tuple: temporary directory containing the script file and path to the file
    """
    quoted_scenarios = (f"'{s}'" for s in scenarios)
    template_path = Path(__file__).parent / "project_modification_script_template.py"
    with open(template_path) as template_file:
        template = template_file.read()
    script_contents = template.format(scenarios=", ".join(quoted_scenarios))
    temp_dir = TemporaryDirectory()  # pylint: disable=consider-using-with
    script_path = WORK_ROOT / _MOD_SCRIPT_NAME
    # TODO: Once this works reliably, we should create the mod_script.py into temp_dir
    # script_path = Path(temp_dir.name) / _MOD_SCRIPT_NAME
    with open(script_path, "w", encoding="utf-8") as script_file:
        script_file.write(script_contents)
    return temp_dir, script_path


def validate_input_data_path(p):
    if not os.path.exists(p):
        return {"success": False, "error": "Path does not exist"}
    if "modelspec.xlsx" in os.listdir(p):
        return {"success": True}
    return {"success": False, "error": "Path is not a valid SiteOpt Data path."}


def validate_project_path(p):
    if not os.path.exists(p):
        return {"success": False, "error": "Path does not exist"}
    if ".spinetoolbox" in os.listdir(p):
        return {"success": True}
    return {"success": False, "error": "Path does not contain a Spine Toolbox project."}


def make_work_folder(client_id, work_folder_name, use_example_data=False):
    configs = get_client_config(client_id) or {}
    config_fpath = get_config_file_dir(client_id) / CONFIG_FILE
    if use_example_data:
        idp = get_example_input_data_path()
    else:
        idp = get_input_data_path()
    pdp = get_project_data_path()
    if not validate_input_data_path(idp)["success"]:
        return {"success": False, "error": f"Bundled input data is invalid: '{idp}'"}
    if not validate_project_path(pdp)["success"]:
        return {"success": False, "error": f"Bundled project data is invalid: '{pdp}'"}
    client_root = Path(get_client_work_root(client_id))
    work_dir = (client_root / work_folder_name).resolve()
    print(f"Creating project to {work_dir}")
    try:
        make_dir(str(work_dir))
        ignored = (".git", ".idea", ".venv", ".gitignore", ".gitkeep", "*.md", "*.png", "*.yaml", "*.yml", "*.css", ".github", "docs")
        shutil.copytree(pdp, str(work_dir), dirs_exist_ok=True, ignore=shutil.ignore_patterns(*ignored))
        shutil.copytree(idp, str(work_dir / "current_input"), dirs_exist_ok=True, ignore=shutil.ignore_patterns(*ignored))
    except OSError as e:
        return {"success": False, "error": f"[OSError] [{e}] Creating work dir failed"}
    work_folders = configs["work_folders"]
    if work_folder_name not in work_folders:
        work_folders[work_folder_name] = str(work_dir)
    edit_config_file(config_fpath, {"work_folders": work_folders})
    return {"success": True, "data": {}}


def make_test_work_folder(client_id, work_folder_name):
    config_fpath = get_config_file_dir(client_id) / CONFIG_FILE
    configs = get_client_config(client_id) or {}
    pdp = get_test_project_data_path()
    if not validate_project_path(pdp)["success"]:
        return {"success": False, "error": f"Bundled project data is invalid: '{pdp}'"}
    client_root = Path(get_client_work_root(client_id))
    work_dir = (client_root / work_folder_name).resolve()
    try:
        make_dir(str(work_dir))
        ignored = (".git", ".idea", ".venv", ".gitignore", ".gitkeep", "*.md", "*.png", "*.yaml", "*.yml", "*.css", ".github", "docs")
        shutil.copytree(pdp, str(work_dir), dirs_exist_ok=True, ignore=shutil.ignore_patterns(*ignored))
    except OSError as e:
        return {"success": False, "error": f"[OSError] [{e}] Creating test work dir failed"}
    work_folders = configs["work_folders"]
    if work_folder_name not in work_folders:
        work_folders[work_folder_name] = str(work_dir)
    edit_config_file(config_fpath, {"work_folders": work_folders})
    return {"success": True, "data": {}}


def build_tree(path, exclude_dirs=None):
    exclude_dirs = set(os.path.abspath(d) for d in (exclude_dirs or []))
    tree = {"children": []}
    entries = os.listdir(path)
    # Sort: files first, then directories
    entries.sort(key=lambda e: os.path.isdir(os.path.join(path, e)))
    for entry in entries:
        full_path = os.path.join(path, entry)
        abs_path = os.path.abspath(full_path)
        # Skip excluded directories
        if abs_path in exclude_dirs:
            continue
        if os.path.isdir(full_path):
            tree["children"].append({
                "name": entry,
                "children": build_tree(full_path, exclude_dirs)["children"]
            })
        else:
            tree["children"].append({"name": entry})
    return tree


def fetch_work_folders_tree(request):
    """Builds and returns the directory tree of available work (project) folders in current context.
    i.e. When in container, host work folders are not shown. When running on host, container work
    folders are not shown."""
    client_id = request.COOKIES.get("client_id") or request.headers.get("X-Client-ID")
    print(f"Client {client_id} is fetching work folder files")
    config_d = get_client_config(client_id)
    # config_d = read_config_file(config.config_path)
    work_folders_dict = config_d.get("work_folders", {})
    trees = list()
    for name, p in work_folders_dict.items():
        if not p.startswith(str(WORK_ROOT)):
            # Skip folders not in current context
            continue
        if not os.path.exists(p):
            print(f"Skipping {p}. Doesn't exist.")
            continue
        excluded_dirs = [os.path.join(p, ".git")]
        tree = build_tree(p, excluded_dirs)
        base_path, dirname = os.path.split(p)
        tree["name"] = dirname  # same as 'name'
        tree["path"] = base_path
        trees.append(tree)
    return JsonResponse({"success": True, "data": trees})


def fetch_work_folder(request, folder_name):
    """Returns the directory tree of a single work folder."""
    print(f"Returning directory tree of {folder_name}")
    client_id = request.COOKIES.get("client_id") or request.headers.get("X-Client-ID")
    config_d = get_client_config(client_id)
    # config_d = read_config_file(config.config_path)
    work_folders_dict = config_d.get("work_folders", {})
    p = work_folders_dict.get(folder_name)
    if not os.path.exists(p):
        return JsonResponse({"success": False, "error": f"Updating project {folder_name} failed"})
    excluded_dirs = [os.path.join(p, ".git")]
    tree = build_tree(p, excluded_dirs)
    base_path, dirname = os.path.split(p)
    tree["name"] = dirname  # same as 'name'
    tree["path"] = base_path
    return JsonResponse({"success": True, "data": tree})


def fetch_current_input_folder(request, folder_name):
    """Returns the files under current_input folder of a given folder (project) name.
    The returned list format is such that it can be used directly in a frontend Toolbar template.
    """
    print(f"Returning current_input of {folder_name}")
    client_id = request.COOKIES.get("client_id") or request.headers.get("X-Client-ID")
    config_d = get_client_config(client_id)
    work_folders_dict = config_d.get("work_folders", {})
    p = work_folders_dict.get(folder_name)
    if not os.path.exists(p):
        return JsonResponse({"success": False, "error": f"Project folder {p} does not exist"})
    input_path = os.path.join(p, "current_input")
    # folders = ["", "connections", "demand", "nodes", "other_units", "production", "representative_periods", "storages"]
    folders = [f for f in os.listdir(input_path) if os.path.isdir(os.path.join(input_path, f))]
    folders = [""] + folders  # Add root folder
    categories = []
    # categories: [{name: "Basic", value: "", options: [{label: scenarios.xlsx, value: scenarios.xlsx}, {}]}, ...]
    for folder in folders:
        entry = {}
        p = os.path.join(input_path, folder)
        if not folder:
            entry["name"] = "Basic"
            entry["value"] = ""
        else:
            entry["name"] = folder.replace("_", " ").capitalize()
            entry["value"] = folder
        entry["options"] = []
        for filename in os.listdir(p):
            if os.path.isfile(os.path.join(p, filename)):
                entry["options"].append({"label": filename, "value": filename})
        categories.append(entry)
    return JsonResponse({"success": True, "data": categories})


def fetch_data(fpath):
    if not os.path.exists(fpath):
        return {"success": False, "error": f"{fpath} does not exist"}
    if fpath.endswith(".xlsx"):
        wb = openpyxl.load_workbook(fpath)
        data = read_excel_as_json(wb)
        return {"success": True, "data": data}
    elif fpath.endswith(".csv"):
        data = read_csv_as_json(fpath)
        return {"success": True, "data": data}
    elif fpath.endswith(".json"):
        try:
            with open(fpath, "r", encoding="utf-8") as fh:
                obj = json.load(fh)
            return {"success": True, "data": {"filetype": "json", "data": obj}}
        except json.JSONDecodeError as e:
            return {"success": False, "error": f"Invalid JSON: {e}"}
        except OSError as e:
            return {"success": False, "error": f"[OSError] {e}"}
    elif fpath.endswith(".md"):
        try:
            with open(fpath, "r", encoding="utf-8") as fh:
                text = fh.read()
            return {"success": True, "data": {"filetype": "md", "data": {"text": text}}}
        except OSError as e:
            return {"success": False, "error": f"[OSError] {e}"}
    else:
        _, ext = os.path.splitext(fpath)
        return {"success": False, "error": f"Reading files with extension '{ext}' not implemented"}


def read_excel_as_json(wb):
    """
    Return:
    {
      "filetype": "xlsx",
      "data": {
        "Sheet1": {
          "columns": [...],
          "rows": [ {col: val, ...}, ... ],
          "validationsByColumn": { "Status": ["A","B"], ... }
        },
        ...
      }
    }
    Assumes row 1 is header.
    """
    out = {"filetype": "xlsx", "data": {}}

    for sheet in wb.worksheets:
        max_col = sheet.max_column
        max_row = sheet.max_row

        # Read header row
        columns = []
        for c in range(1, max_col + 1):
            v = sheet.cell(row=1, column=c).value
            columns.append("" if v is None else str(v))

        # Build rows (row 2..max_row)
        rows = []
        for r in range(2, max_row + 1):
            row_obj = {}
            for c, col_name in enumerate(columns, start=1):
                row_obj[col_name] = sheet.cell(row=r, column=c).value
            rows.append(row_obj)

        validations_by_column = _extract_validations_by_column(wb, sheet, columns)

        out["data"][sheet.title] = {
            "columns": columns,
            "rows": rows,
            "validationsByColumn": validations_by_column,
        }

    return out


def read_csv_as_json(p):
    with open(p, newline="", encoding="utf-8") as fp:
        reader = csv.DictReader(fp)
        rows = list(reader)
        cols = reader.fieldnames or []
    return {"filetype": "csv", "data": {"columns": cols, "rows": rows}}


def make_dir(p):
    """Creates a directory if it doesn't exist.

    Args:
        p: Absolute path to wanted dir

    Raises:
        OSError if operation failed.
    """
    os.makedirs(p, exist_ok=True)


def make_config_file(p):
    """Creates a dummy config file.

    Args:
        p (str): Full path to config file
    """
    d = {"input_data_path": get_input_data_path(),
         "project_data_path": get_project_data_path(),
         "test_project_data_path": get_test_project_data_path(),
         "work_folders": {}}
    with open(p, "w") as fp:
        json.dump(d, fp, indent=4)


def edit_config_file(p, new_key_value):
    print(f"new_key_value:{new_key_value}")
    config_d = read_config_file(p)
    config_d.update(new_key_value)
    with open(p, "w") as fp:
        json.dump(config_d, fp, indent=4)


def read_config_file(p):
    """Reads a file from given path and returns the contents as a dict.

    Args:
        p (str): Full path to config file
    """
    try:
        with open(p, "r") as fh:
            try:
                config_dict = json.load(fh)
            except json.decoder.JSONDecodeError:
                print(f"[JSONDecodeError] in config file {p}. Invalid JSON, maybe?")
                return {}
    except OSError:
        print(f"[OSError] Config file {p} missing, maybe?")
        return {}
    return config_dict


def is_path_inside_any_work_folder(config_fpath: str, target_path: str) -> bool:
    cfg = read_config_file(config_fpath)
    work_folders = cfg.get("work_folders", {})
    target = os.path.abspath(target_path)

    for wf in work_folders.values():
        wf_abs = os.path.abspath(wf)
        try:
            common = os.path.commonpath([target, wf_abs])
        except ValueError:
            continue
        if common == wf_abs:
            return True
    return False


def _save_md(fpath: str, data, meta: dict):
    # data can be either raw string or {"text": "..."} – allow both for flexibility
    if isinstance(data, dict):
        text = data.get("text", "")
    else:
        text = "" if data is None else str(data)

    if not fpath.endswith(".md"):
        return {"success": False, "error": "Not a .md file."}

    with open(fpath, "w", encoding="utf-8", newline="\n") as fp:
        fp.write(text)

    return {"success": True}


def _save_csv(fpath: str, data, meta: dict):
    if not isinstance(data, list):
        return {"success": False, "error": "csv save expects a list of row objects."}
    if len(data) == 0:
        return {"success": False, "error": "No data to save."}

    columns = meta.get("columns")
    if not columns:
        cols_set = set()
        for row in data:
            if isinstance(row, dict):
                cols_set.update(row.keys())
        columns = list(cols_set)

    with open(fpath, "w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in data:
            if not isinstance(row, dict):
                return {"success": False, "error": "Each CSV row must be an object/dict."}
            writer.writerow({c: row.get(c, "") for c in columns})

    return {"success": True}


def _save_json(fpath: str, data, meta: dict):
    if not fpath.endswith(".json"):
        return {"success": False, "error": "Not a .json file."}
    obj = None

    if isinstance(data, dict) and "text" in data:
        try:
            obj = json.loads(data["text"])
        except json.JSONDecodeError as e:
            return {"success": False, "error": f"Invalid JSON: {e}"}
    else:
        obj = data

    try:
        with open(fpath, "w", encoding="utf-8", newline="\n") as fp:
            json.dump(obj, fp, ensure_ascii=False, indent=2)
            fp.write("\n")
        return {"success": True}
    except TypeError as e:
        return {"success": False, "error": f"JSON is not serializable: {e}"}


def _save_xlsx(fpath: str, data, meta: dict):
    """
    Expects:
      data = dictionary containing the whole workbook
      meta = {"sheet": "SheetName", "columns": ["A","B",...]}
    """
    if not fpath.endswith(".xlsx"):
        return {"success": False, "error": "Not an .xlsx file."}
    if not isinstance(data, dict):
        return {"success": False, "error": "xlsx save expects a dictionary containing sheets and wb data."}
    wb = openpyxl.load_workbook(fpath)
    # Loop all sheets
    for sheet_name in data.keys():
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} not found in wb {fpath}. Skipping...")
            continue
        ws = wb[sheet_name]
        columns = data[sheet_name]["columns"]
        rows = data[sheet_name]["rows"]
        # print(f"Processing sheet {sheet_name}. Columns: {columns}")
        # print(f"rows:{rows}")
        ws.delete_rows(1, ws.max_row)
        # Make header
        for c, col in enumerate(columns, start=1):
            ws.cell(row=1, column=c, value=col)
        # Add rows
        for r_i, row in enumerate(rows, start=2):
            if not isinstance(row, dict):
                return {"success": False, "error": "Each XLSX row must be an object/dict."}
            for c_i, col in enumerate(columns, start=1):
                ws.cell(row=r_i, column=c_i, value=row.get(col, ""))
    # Save workbook
    wb.save(fpath)
    return {"success": True}


def save_file(client_id: str, fpath: str, filetype: str, data, meta: dict):
    if not fpath or not filetype:
        return {"success": False, "error": "Missing 'path' or 'filetype'."}
    config_fpath = get_config_file_dir(client_id) / CONFIG_FILE
    if not is_path_inside_any_work_folder(str(config_fpath), fpath):
        return {"success": False, "error": "Refusing to write outside work folders."}

    if not os.path.exists(fpath):
        return {"success": False, "error": f"File does not exist: {fpath}"}

    save_handlers = {
        "md": _save_md,
        "json": _save_json,
        "csv": _save_csv,
        "xlsx": _save_xlsx,
    }

    handler = save_handlers.get(filetype)
    if not handler:
        return {"success": False, "error": f"Saving not implemented for '{filetype}'."}

    try:
        return handler(fpath, data, meta)
    except OSError as e:
        return {"success": False, "error": f"[OSError] {e}"}
    except Exception as e:
        return {"success": False, "error": f"[Error] {e}"}
    

def _parse_list_formula(formula1):
    """
    Returns either:
      - list[str] for explicit lists: '"A,B,C"'
      - {"range": "Sheet2!$A$1:$A$5"} for range-based lists: "=Sheet2!$A$1:$A$5"
      - None if not recognized
    """
    if not formula1:
        return None

    f = str(formula1).strip()

    if f.startswith('"') and f.endswith('"'):
        inner = f[1:-1]
        sep = "," if "," in inner else ";"
        return [x.strip() for x in inner.split(sep) if x.strip()]

    if f.startswith("="):
        return {"range": f[1:]}

    return None


def _resolve_range_list(wb, range_expr):
    """
    range_expr examples:
      Sheet2!$A$1:$A$5
      'Sheet 2'!$A$1:$A$5
    Returns list[str] or None.
    """
    if "!" not in range_expr:
        return None

    sheet_part, addr = range_expr.split("!", 1)
    sheet_name = sheet_part.strip("'").strip()

    if sheet_name not in wb.sheetnames:
        return None

    ws2 = wb[sheet_name]
    values = []

    try:
        cells = ws2[addr]
    except Exception:
        return None

    for row in cells:
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            if s != "":
                values.append(s)

    out = []
    seen = set()
    for x in values:
        if x not in seen:
            out.append(x)
            seen.add(x)
    return out


def _extract_validations_by_column(wb, ws, columns):
    """
    Builds:
      { "ColumnHeader": ["opt1","opt2",...], ... }
    Only handles dv.type == "list" for now (dropdowns).
    Works best when validation applies to an entire column range.
    """
    validations = {}

    dvs = getattr(ws, "data_validations", None)
    if not dvs or not getattr(dvs, "dataValidation", None):
        return validations

    def header_for_col_index(ci):
        if 1 <= ci <= len(columns):
            return columns[ci - 1]
        return None

    for dv in dvs.dataValidation:
        if getattr(dv, "type", None) != "list":
            continue

        parsed = _parse_list_formula(getattr(dv, "formula1", None))
        options = None

        if isinstance(parsed, list):
            options = parsed
        elif isinstance(parsed, dict) and "range" in parsed:
            options = _resolve_range_list(wb, parsed["range"])

        if not options:
            continue

        for cell_range in dv.sqref.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))

            for col_idx in range(min_col, max_col + 1):
                header = header_for_col_index(col_idx)
                if not header:
                    continue

                if header not in validations:
                    validations[header] = options

    return validations
